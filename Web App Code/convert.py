import csv
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook(filename='dynamodb_data.xlsx')

# Select the first sheet
ws = wb.active

# Add annotations
ws.insert_rows(0, 2)  # Insert 2 rows at the top
ws['A1'] = '# DDL'
ws['A2'] = '# DML'

# Save the workbook as a CSV file
with open('annotated_file.csv', 'w', newline="") as f:
    c = csv.writer(f)
    for r in ws.rows:
        c.writerow([cell.value for cell in r])